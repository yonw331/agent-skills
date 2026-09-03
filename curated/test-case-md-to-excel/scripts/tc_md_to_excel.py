"""
tc_md_to_excel.py - Markdown 测试用例批量转 Excel 工具

功能：
1. 批量解析指定目录下的 Markdown 测试用例文件
2. 将所有测试用例合并到一个 Excel 文件中
3. 按模块分组，自动生成编号（格式：TC-<模块拼音首字母>-序号）

首选 Markdown 格式（详见 assets/case_template.md）：
  # 所属模块
  ## 功能点
  ### 用例标题
  - [ ] 1️⃣  前置条件 | 操作步骤 | 预期结果

兼容旧格式：
  - [Px] 前置条件 | 操作步骤 | 预期结果 | 执行结果

用法：
    python3 tc_md_to_excel.py [测试用例目录] [--prefix 前缀] [--output 输出路径]

示例：
    python3 tc_md_to_excel.py ./testcases
    python3 tc_md_to_excel.py ./testcases --prefix TC
    python3 tc_md_to_excel.py ./testcases --output ./result/用例.xlsx
"""

import argparse
import logging
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# 依赖检查（所有外部依赖集中在此）
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误：需要安装 openpyxl 库")
    print("请运行：python3 -m pip install openpyxl")
    sys.exit(1)

try:
    from pypinyin import lazy_pinyin
except ImportError:
    print("错误：需要安装 pypinyin 库")
    print("请运行：python3 -m pip install pypinyin")
    sys.exit(1)

# ---------------------------------------------------------------------------
# 日志配置
# ---------------------------------------------------------------------------
LOG = logging.getLogger("tc_md_to_excel")
LOG.setLevel(logging.DEBUG)
_handler = logging.StreamHandler(sys.stdout)
_handler.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
LOG.addHandler(_handler)

# ---------------------------------------------------------------------------
# Excel 列定义（9 列）
# ---------------------------------------------------------------------------
EXCEL_HEADERS = [
    "用例编号",   # A - TC-WYKMYK-001
    "所属模块",   # B
    "功能点",     # C
    "优先级",     # D
    "用例标题",   # E
    "前置条件",   # F
    "操作步骤",   # G
    "预期结果",   # H
    "执行结果",   # I
]

COLUMN_COUNT = len(EXCEL_HEADERS)

# 优先级 -> 背景色
PRIORITY_COLORS = {
    "P0": "FF6B6B",
    "P1": "FFD93D",
    "P2": "6BCB77",
    "P3": "95A5A6",
}

TASK_PRIORITY_MAP = {
    "1️⃣": "P0",
    "2️⃣": "P1",
    "3️⃣": "P2",
    "4️⃣": "P3",
}

# 测试结果 -> 背景色
RESULT_COLORS = {
    "未执行":   "D9E1F2",
    "已执行":   "5B9BD5",
    "测试通过": "6BCB77",
    "未通过":   "FF6B6B",
}

# 列宽配置 {列号: 宽度}
COLUMN_WIDTHS = {
    1: 22,  # 用例编号
    2: 28,  # 所属模块
    3: 22,  # 功能点
    4: 10,  # 优先级
    5: 36,  # 用例标题
    6: 36,  # 前置条件
    7: 42,  # 操作步骤
    8: 42,  # 预期结果
    9: 18,  # 执行结果
}

# 文件名匹配模式（小写）
FILENAME_PATTERNS = [
    "testcase", "测试用例", "用例", "需求确认",
    "spec-", "tc-", "test_case", "case-",
]


# =========================================================================
# 数据结构
# =========================================================================

@dataclass
class TestCase:
    """单条测试用例"""
    module: str = ""          # 所属模块（# 标题）
    feature: str = ""         # 功能点（## 标题）
    priority: str = ""        # 优先级 P0-P3
    title: str = ""           # 用例标题（### 标题）
    precondition: str = ""    # 前置条件
    steps: str = ""           # 操作步骤
    expected: str = ""        # 预期结果
    test_result: str = ""     # 执行结果
    source_file: str = ""     # 来源文件名（调试用）
    case_no: str = ""         # 生成的用例编号


# =========================================================================
# 工具函数
# =========================================================================

def get_module_abbr(module_name: str) -> str:
    """
    从中文模块名提取拼音首字母，转大写。

    纯英文/数字模块名直接返回大写。
    空字符串返回 "M"（兜底）。
    """
    if not module_name:
        return "M"

    # 分离中文字符和非中文字符
    chinese_chars = re.findall(r"[\u4e00-\u9fff]", module_name)
    non_chinese = re.sub(r"[\u4e00-\u9fff]", "", module_name).strip()

    if chinese_chars:
        try:
            initials = lazy_pinyin("".join(chinese_chars))
            abbr = "".join(p[0].upper() for p in initials if p)
        except Exception:
            abbr = ""
    else:
        abbr = ""

    if non_chinese:
        # 英文/数字部分保留大写，去掉空格和特殊字符
        clean = re.sub(r"[^a-zA-Z0-9]", "", non_chinese).upper()
        abbr += clean

    return abbr if abbr else "M"


def format_content_with_numbering(content: str) -> str:
    """
    将分号分隔的多条内容转为编号列表。

    - 分号 ; 或 ；，以及解析阶段保留的换行，作为分隔符
    - 如果源数据已有编号（如 "1." / "1、"），保留原样
    - 单条内容不添加编号
    """
    if not content:
        return content

    # 解析器会先将分号转换为换行，因此两种分隔符都要支持。
    parts = re.split(r"[;；\n]+", content)
    parts = [p.strip() for p in parts]

    # 过滤空项
    parts = [p for p in parts if p]

    if not parts:
        return content

    # 单条直接返回
    if len(parts) == 1:
        return parts[0]

    # 检查是否已有编号
    has_numbering = bool(re.match(r"^\d+[.、)\]：:]\s*", parts[0]))

    if has_numbering:
        return "\n".join(parts)

    return "\n".join(f"{i}.{p}" for i, p in enumerate(parts, 1))


def semicolon_to_newline(text: str) -> str:
    """将分号分隔的内容转为换行（用于内部存储）。"""
    if not text:
        return ""
    return re.sub(r"[;；]\s*", "\n", text).strip()


def parse_test_result(result_str: str) -> str:
    """
    解析执行结果字段。

    格式示例：
      - [x] 通过 - [ ] 未通过           → "测试通过"
      - [ ] 通过 - [x] 未通过           → "未通过"
      - [ ] 通过 - [x] 未通过 登录失败   → "未通过：登录失败"
      - (空/无匹配)                     → "--"
    """
    if not result_str:
        return "--"

    result_str = result_str.strip()

    # 正则：捕获 "通过" 勾选、"未通过" 勾选、备注内容 三个组
    pattern = r"-\s*\[(x| )\]\s*通过\s*-\s*\[(x| )\]\s*未通过(.*)"
    match = re.search(pattern, result_str)

    if not match:
        # 兜底：检查是否包含关键文字
        if "通过" in result_str and "未通过" not in result_str:
            return "测试通过"
        return "--"

    passed_checked = match.group(1) == "x"
    failed_checked = match.group(2) == "x"
    remark = match.group(3).strip()

    if passed_checked:
        return "测试通过"

    if failed_checked:
        return f"未通过：{remark}" if remark else "未通过"

    return "--"


# =========================================================================
# Markdown 解析器
# =========================================================================

class MarkdownParser:
    """
    解析符合模板格式的 Markdown 测试用例文件。

    状态机逐行解析，维护 module / feature / title 上下文，
    遇到 Obsidian 三段式或旧版 `[Px]` 四段式行时提取一条完整用例。
    """

    # 头部需要跳过的行前缀
    SKIP_PREFIXES = (
        "---", "title:", "author:", "priority:", "status:",
        "created:", "tags:", "source:", "date:", "updated:",
    )

    def __init__(self, content: str, source_file: str = ""):
        self.content = content
        self.lines = content.split("\n")
        self.source_file = source_file

        # 解析状态
        self.current_module = ""
        self.current_feature = ""
        self.current_title = ""

        # 结果
        self.cases: List[TestCase] = []

    def _is_skip_line(self, line: str) -> bool:
        """判断是否应跳过该行（头部元信息 / 空行 / 引用块）。"""
        stripped = line.strip()
        if not stripped:
            return True
        if stripped.startswith(">"):
            return True
        for prefix in self.SKIP_PREFIXES:
            if stripped.startswith(prefix):
                return True
        # Markdown 表格分隔行
        if stripped.startswith("|") and re.match(r"^\|[\s\-:|]+\|$", stripped):
            return True
        return False

    def _is_in_header_zone(self) -> bool:
        """
        是否仍处于头部区域。
        头部区域：从文件开头到遇到第一个 `# ` 标题为止。
        """
        return len(self.cases) == 0 and self.current_module == ""

    def _parse_module(self, line: str) -> Optional[str]:
        """解析 `# 模块名`，返回模块名或 None。"""
        match = re.match(r"^#\s+(.+)$", line)
        if not match:
            return None

        module_name = match.group(1).strip()
        # 去掉括号注释，如 "模块名称（WYKMYK）"
        module_name = re.sub(r"\s*[（(].*?[）)]\s*$", "", module_name).strip()
        return module_name if module_name else None

    def _parse_feature(self, line: str) -> Optional[str]:
        """解析 `## 功能点`，返回功能点名称或 None。"""
        match = re.match(r"^##\s+(.+)$", line)
        if match:
            return match.group(1).strip() or None
        return None

    def _parse_title(self, line: str) -> Optional[str]:
        """解析 `### 用例标题`，返回标题或 None。"""
        match = re.match(r"^###\s+(.+)$", line)
        if match:
            return match.group(1).strip() or None
        return None

    def _parse_case_line(self, line: str) -> Optional[TestCase]:
        """
        解析首选格式与兼容格式：
        - `- [ ] 1️⃣ 前置条件 | 操作步骤 | 预期结果`
        - `- [Px] 前置条件 | 操作步骤 | 预期结果 | 执行结果`

        返回 TestCase 或 None（非用例行）。
        """
        stripped = line.strip()

        task_match = re.match(
            r"-\s*\[([ xX])\]\s*(1️⃣|2️⃣|3️⃣|4️⃣)\s+(.+)",
            stripped,
            re.DOTALL,
        )
        if task_match:
            checked, priority_icon, body = task_match.groups()
            parts = [part.strip() for part in body.split("|", 2)]
            while len(parts) < 3:
                parts.append("")
            precondition, steps, expected = parts

            return TestCase(
                module=self.current_module,
                feature=self.current_feature,
                priority=TASK_PRIORITY_MAP[priority_icon],
                title=self.current_title,
                precondition=semicolon_to_newline(precondition),
                steps=semicolon_to_newline(steps),
                expected=semicolon_to_newline(expected),
                test_result="已执行" if checked.lower() == "x" else "未执行",
                source_file=self.source_file,
            )

        legacy_task_match = re.match(
            r"-\s*\[([ xX])\]\s*\[(P[0-3])\](?:\[[^\]]+\])*\s*(.+)",
            stripped,
            re.DOTALL,
        )
        if legacy_task_match:
            checked, priority, body = legacy_task_match.groups()
            parts = [part.strip() for part in body.split("|", 2)]
            while len(parts) < 3:
                parts.append("")
            precondition, steps, expected = parts

            return TestCase(
                module=self.current_module,
                feature=self.current_feature,
                priority=priority,
                title=self.current_title,
                precondition=semicolon_to_newline(precondition),
                steps=semicolon_to_newline(steps),
                expected=semicolon_to_newline(expected),
                test_result="已执行" if checked.lower() == "x" else "未执行",
                source_file=self.source_file,
            )

        priority_match = re.match(r"-\s*\[(P\d)\]\s*(.+)", stripped, re.DOTALL)
        if not priority_match:
            return None

        priority = priority_match.group(1)
        body = priority_match.group(2).strip()

        parts = [p.strip() for p in body.split("|", 3)]

        # 补齐到 4 段
        while len(parts) < 4:
            parts.append("")

        precondition, steps, expected, result_raw = parts[0], parts[1], parts[2], parts[3]

        return TestCase(
            module=self.current_module,
            feature=self.current_feature,
            priority=priority,
            title=self.current_title,
            precondition=semicolon_to_newline(precondition),
            steps=semicolon_to_newline(steps),
            expected=semicolon_to_newline(expected),
            test_result=parse_test_result(result_raw),
            source_file=self.source_file,
        )

    def parse(self) -> List[TestCase]:
        """执行解析，返回用例列表。"""
        for line in self.lines:
            stripped = line.strip()

            # 头部区域跳过
            if self._is_in_header_zone() and self._is_skip_line(line):
                continue

            # 尝试匹配各级标题
            if stripped.startswith("### "):
                title = self._parse_title(stripped)
                if title is not None:
                    self.current_title = title
                continue

            if stripped.startswith("## ") and not stripped.startswith("### "):
                feature = self._parse_feature(stripped)
                if feature is not None:
                    self.current_feature = feature
                    self.current_title = ""  # 进入新功能点，重置标题
                continue

            if stripped.startswith("# ") and not stripped.startswith("## "):
                module = self._parse_module(stripped)
                if module is not None:
                    self.current_module = module
                    self.current_feature = ""
                    self.current_title = ""
                continue

            # 尝试解析用例行
            case = self._parse_case_line(stripped)
            if case is not None:
                self.cases.append(case)

        return self.cases


# =========================================================================
# 用例编号生成器
# =========================================================================

def generate_case_numbers(cases: List[TestCase], prefix: str = "TC") -> None:
    """
    为用例列表生成编号，直接修改 case.case_no。

    格式：{prefix}-{模块拼音首字母}-{三位序号}
    示例：TC-WYKMYK-001

    编号按模块分组，每个模块独立计数。
    """
    module_counter: Dict[str, int] = {}
    module_abbr_cache: Dict[str, str] = {}

    for case in cases:
        module_key = case.module if case.module else "未知模块"

        # 缓存模块缩写，避免重复计算
        if module_key not in module_abbr_cache:
            module_abbr_cache[module_key] = get_module_abbr(module_key)

        module_counter[module_key] = module_counter.get(module_key, 0) + 1

        case.case_no = (
            f"{prefix}-{module_abbr_cache[module_key]}-{module_counter[module_key]:03d}"
        )


# =========================================================================
# 文件扫描
# =========================================================================

def scan_md_files(directory: str) -> List[str]:
    """
    扫描目录下的 Markdown 测试用例文件。

    匹配规则（文件名包含以下关键词之一）：
      testcase, 测试用例, 用例, 需求确认, spec-, tc-, test_case
    """
    md_files: List[str] = []

    if not os.path.isdir(directory):
        LOG.error("目录不存在: %s", directory)
        return md_files

    for entry in sorted(os.listdir(directory)):
        if not entry.lower().endswith(".md"):
            continue

        entry_lower = entry.lower()
        matched = any(
            entry_lower.startswith(pat) or pat in entry_lower
            for pat in FILENAME_PATTERNS
        )
        if matched:
            md_files.append(os.path.join(directory, entry))

    return md_files


# =========================================================================
# Excel 生成器
# =========================================================================

class ExcelGenerator:
    """将 TestCase 列表写入格式化的 Excel 文件。"""

    def __init__(self, cases: List[TestCase]):
        self.cases = cases
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "测试用例"

    def generate(self, output_path: str) -> None:
        """生成并保存 Excel 文件。"""
        self._setup_header()
        self._fill_data()
        self._apply_styles()
        self._adjust_column_widths()
        self.wb.save(output_path)

    # -- 内部方法 --

    def _setup_header(self) -> None:
        """写入表头行。"""
        for col_idx, header_text in enumerate(EXCEL_HEADERS, 1):
            cell = self.ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    def _fill_data(self) -> None:
        """填充用例数据。"""
        for row_idx, case in enumerate(self.cases, 2):
            self.ws.cell(row=row_idx, column=1, value=case.case_no)
            self.ws.cell(row=row_idx, column=2, value=case.module)
            self.ws.cell(row=row_idx, column=3, value=case.feature)
            self.ws.cell(row=row_idx, column=4, value=case.priority)
            self.ws.cell(row=row_idx, column=5, value=case.title)
            self.ws.cell(
                row=row_idx, column=6,
                value=format_content_with_numbering(case.precondition),
            )
            self.ws.cell(
                row=row_idx, column=7,
                value=format_content_with_numbering(case.steps),
            )
            self.ws.cell(
                row=row_idx, column=8,
                value=format_content_with_numbering(case.expected),
            )
            self.ws.cell(row=row_idx, column=9, value=case.test_result)

    def _apply_styles(self) -> None:
        """应用单元格样式。"""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row_idx in range(2, len(self.cases) + 2):
            # 优先级列（D = column 4）
            priority_cell = self.ws.cell(row=row_idx, column=4)
            pval = priority_cell.value
            if pval in PRIORITY_COLORS:
                priority_cell.fill = PatternFill(
                    start_color=PRIORITY_COLORS[pval],
                    end_color=PRIORITY_COLORS[pval],
                    fill_type="solid",
                )
                priority_cell.font = Font(bold=True)
                priority_cell.alignment = Alignment(horizontal="center", vertical="center")

            # 测试结果列（I = column 9）
            result_cell = self.ws.cell(row=row_idx, column=9)
            rval = result_cell.value
            if rval and rval.startswith("未通过"):
                result_cell.fill = PatternFill(
                    start_color=RESULT_COLORS["未通过"],
                    end_color=RESULT_COLORS["未通过"],
                    fill_type="solid",
                )
                result_cell.font = Font(bold=True)
            elif rval in RESULT_COLORS:
                result_cell.fill = PatternFill(
                    start_color=RESULT_COLORS[rval],
                    end_color=RESULT_COLORS[rval],
                    fill_type="solid",
                )
                result_cell.font = Font(bold=True)

            # 所有数据列：边框 + 自动换行
            for col_idx in range(1, COLUMN_COUNT + 1):
                cell = self.ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if cell.alignment == Alignment():
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _adjust_column_widths(self) -> None:
        """设置列宽和行高。"""
        for col_idx, width in COLUMN_WIDTHS.items():
            letter = get_column_letter(col_idx)
            self.ws.column_dimensions[letter].width = width

        # 表头行高
        self.ws.row_dimensions[1].height = 30

        # 数据行高（内容多时自动撑开，这里设最小值）
        for row_idx in range(2, len(self.cases) + 2):
            self.ws.row_dimensions[row_idx].height = 60

        # 冻结首行
        self.ws.freeze_panes = "A2"

        # 自动筛选
        self.ws.auto_filter.ref = self.ws.dimensions


# =========================================================================
# 统计打印
# =========================================================================

def print_stats(cases: List[TestCase]) -> None:
    """打印用例统计摘要。"""
    total = len(cases)
    if total == 0:
        LOG.warning("未解析到任何测试用例")
        return

    priority_count: Dict[str, int] = {}
    module_count: Dict[str, int] = {}

    for case in cases:
        priority_count[case.priority] = priority_count.get(case.priority, 0) + 1
        if case.module:
            module_count[case.module] = module_count.get(case.module, 0) + 1

    LOG.info("总计: %d 条测试用例", total)

    for level in ("P0", "P1", "P2", "P3"):
        label = {"P0": "必测", "P1": "重要", "P2": "一般", "P3": "可选"}[level]
        LOG.info("  %s（%s）：%d 条", level, label, priority_count.get(level, 0))

    if module_count:
        LOG.info("模块分布:")
        for module, count in sorted(module_count.items(), key=lambda x: x[1], reverse=True):
            LOG.info("  - %s: %d 条", module, count)


# =========================================================================
# 主函数
# =========================================================================

def build_default_output_path(case_dir: str) -> str:
    """构建默认输出路径：{case_dir}/excel-case/{目录名}_测试用例_{时间戳}.xlsx"""
    output_dir = os.path.join(case_dir, "excel-case")
    os.makedirs(output_dir, exist_ok=True)

    dir_name = os.path.basename(os.path.normpath(case_dir))
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(output_dir, f"{dir_name}_测试用例_{timestamp}.xlsx")


def main() -> None:
    """入口函数。"""
    parser = argparse.ArgumentParser(
        description="Markdown 测试用例批量转 Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例：
  python3 tc_md_to_excel.py ./testcases
  python3 tc_md_to_excel.py ./testcases --prefix TC
  python3 tc_md_to_excel.py ./testcases --output ./result.xlsx
        """,
    )
    parser.add_argument(
        "case_dir",
        nargs="?",
        default=".",
        help="测试用例目录（默认当前目录）",
    )
    parser.add_argument(
        "--prefix", "-p",
        default="TC",
        help="用例编号前缀（默认 TC）",
    )
    parser.add_argument(
        "--output", "-o",
        default="",
        help="输出 Excel 路径（默认自动生成到 {case_dir}/excel-case/）",
    )

    args = parser.parse_args()

    # 1. 验证输入目录
    case_dir = os.path.abspath(args.case_dir)
    if not os.path.isdir(case_dir):
        LOG.error("目录不存在: %s", case_dir)
        sys.exit(1)

    # 2. 扫描 MD 文件
    md_files = scan_md_files(case_dir)
    if not md_files:
        LOG.warning("在 %s 中未找到测试用例文件", case_dir)
        sys.exit(1)

    LOG.info("找到 %d 个测试用例文件:", len(md_files))
    for f in md_files:
        LOG.info("  - %s", os.path.basename(f))

    # 3. 解析所有文件
    all_cases: List[TestCase] = []
    for md_path in md_files:
        LOG.info("解析: %s", os.path.basename(md_path))
        try:
            with open(md_path, "r", encoding="utf-8") as fh:
                content = fh.read()
        except UnicodeDecodeError:
            LOG.warning("  文件编码异常，尝试 GBK: %s", os.path.basename(md_path))
            try:
                with open(md_path, "r", encoding="gbk") as fh:
                    content = fh.read()
            except Exception as exc:
                LOG.error("  读取失败: %s", exc)
                continue
        except Exception as exc:
            LOG.error("  读取失败: %s", exc)
            continue

        parser = MarkdownParser(content, source_file=os.path.basename(md_path))
        cases = parser.parse()
        all_cases.extend(cases)
        LOG.info("  提取 %d 条用例", len(cases))

    if not all_cases:
        LOG.error("所有文件中均未解析到测试用例")
        sys.exit(1)

    # 4. 打印统计
    print_stats(all_cases)

    # 5. 生成编号
    generate_case_numbers(all_cases, prefix=args.prefix)

    # 6. 生成 Excel
    output_path = args.output if args.output else build_default_output_path(case_dir)
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    LOG.info("生成 Excel: %s", output_path)
    generator = ExcelGenerator(all_cases)
    generator.generate(output_path)

    LOG.info("完成! 共 %d 条用例 -> %s", len(all_cases), output_path)


if __name__ == "__main__":
    main()
